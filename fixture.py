import random, math
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def luckydraw(category):
    if len(category.members) == 0:
        return None
    return random.choice(category.members)

def choose_and_remove(category, team):
    chosen_player = luckydraw(category)

    if chosen_player is None:
        return category, team  # nothing to remove

    team.members = np.append(team.members, chosen_player)
    team.total_points += category.points
    category.members = category.members[category.members != chosen_player]

    return category, team


def create_balanced_teams(players_dict, setters_list, num_teams=4):
    """
    Create balanced teams with constraints:
    - Each team has exactly 1 setter
    - Women distributed equally (should have same number of women per team)
    - Points per team should be balanced
    """
    teams = [Team([], 0) for _ in range(num_teams)]
    
    # Track women count per team
    women_per_team = [0] * num_teams
    
    # Step 1: Distribute setters (1 per team)
    setter_pool = setters_list.copy()
    random.shuffle(setter_pool)
    
    for team_idx, setter_name in enumerate(setter_pool[:num_teams]):
        # Find which tier this setter belongs to and get their points
        setter_points = 0
        is_woman = False
        for tier_name, tier_obj in players_dict.items():
            if setter_name in tier_obj.members:
                setter_points = tier_obj.points
                # Check if this setter is a woman BEFORE removing
                if tier_name in ['womenA', 'womenB']:
                    is_woman = True
                # Remove from tier to avoid duplication
                tier_obj.members = tier_obj.members[tier_obj.members != setter_name]
                break
        
        teams[team_idx].members = np.append(teams[team_idx].members, setter_name)
        teams[team_idx].total_points += setter_points
        
        # Track if setter is a woman
        if is_woman:
            women_per_team[team_idx] += 1
    
    # Step 2: Distribute women equally (prioritize equal distribution)
    women_tiers = ['womenA', 'womenB']
    for tier_name in women_tiers:
        if tier_name not in players_dict:
            continue
        
        tier = players_dict[tier_name]
        while len(tier.members) > 0:
            # Find team with least women
            team_idx = women_per_team.index(min(women_per_team))
            
            # Pick a woman from this tier
            chosen = luckydraw(tier)
            if chosen is None:
                break
            
            teams[team_idx].members = np.append(teams[team_idx].members, chosen)
            teams[team_idx].total_points += tier.points
            tier.members = tier.members[tier.members != chosen]
            women_per_team[team_idx] += 1
    
    # Step 3: Fill remaining spots trying to balance points
    # Sort teams by current points (ascending) to fill lower-point teams first
    remaining_tiers = ['menA', 'menB', 'menC', 'wildcard']
    
    while True:
        # Check if all teams are full (6 players each)
        if all(len(team.members) >= 6 for team in teams):
            break
        
        # Find available tiers
        available_tiers = [tier_name for tier_name in remaining_tiers 
                          if tier_name in players_dict and len(players_dict[tier_name].members) > 0]
        
        if not available_tiers:
            break
        
        # Find team with lowest points that's not full
        team_idx = min([i for i, team in enumerate(teams) if len(team.members) < 6],
                      key=lambda i: teams[i].total_points)
        
        # Try to pick from a tier that balances points
        # Prefer higher-point tiers for lower-point teams
        current_avg_points = sum(t.total_points for t in teams) / num_teams
        
        if teams[team_idx].total_points < current_avg_points:
            # Pick from higher-point tier
            tier_name = max(available_tiers, key=lambda t: players_dict[t].points)
        else:
            # Pick from lower-point tier
            tier_name = min(available_tiers, key=lambda t: players_dict[t].points)
        
        tier = players_dict[tier_name]
        chosen = luckydraw(tier)
        
        if chosen is None:
            continue
        
        teams[team_idx].members = np.append(teams[team_idx].members, chosen)
        teams[team_idx].total_points += tier.points
        tier.members = tier.members[tier.members != chosen]
    
    return teams

class Tier:
    def __init__(self, data, points):
        # Store a NumPy array as an attribute
        self.members = np.array(data, dtype=str)
        self.points = points

class Team:
    def __init__(self, members, total_points):
        self.members = np.array(members, dtype=str)
        self.total_points = total_points

def read_tier_list(excel_file):
    """
    Read tier list from Excel file with format:
    Column A: men names
    Column B: men tier (A/B/C)
    Column D: women names
    Column E: women tier (A/B)
    Column G: setter names
    
    New tiering system:
    - Men A - 4pt
    - Men B - 3pt
    - Women A - 3pt
    - Men C - 2pt
    - Women B - 2pt
    - Wildcard - 1pt
    """
    wb = load_workbook(excel_file)
    ws = wb.active
    
    menA, menB, menC = [], [], []
    womenA, womenB = [], []
    wildcard = []
    setter = []
    
    # Read men (columns A and B)
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value  # Column A
        tier = ws.cell(row=row, column=2).value  # Column B
        if name:
            name = str(name).strip()  # Clean up whitespace
            if tier == 'A':
                menA.append(name)
            elif tier == 'B':
                menB.append(name)
            elif tier == 'C':
                menC.append(name)
            elif tier == 'Wildcard' or tier == 'W':
                wildcard.append(name)
    
    # Read women (columns D and E)
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=4).value  # Column D
        tier = ws.cell(row=row, column=5).value  # Column E
        if name:
            name = str(name).strip()  # Clean up whitespace
            if tier == 'A':
                womenA.append(name)
            elif tier == 'B':
                womenB.append(name)
            elif tier == 'Wildcard' or tier == 'W':
                wildcard.append(name)
    
    # Read setter (column G)
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=7).value  # Column G
        if name:
            name = str(name).strip()  # Clean up whitespace
            setter.append(name)
    
    wb.close()
    
    return {
        'menA': menA,
        'menB': menB,
        'menC': menC,
        'womenA': womenA,
        'womenB': womenB,
        'wildcard': wildcard,
        'setter': setter
    }

def main():
    # Read tier list from Excel file
    tier_list_file = 'tier_list.xlsx'  # Change this to your tier list file name
    tier_data = read_tier_list(tier_list_file)
    
    # Create player info mapping for colors (player_name -> {gender, tier})
    # Setters are also in the tier lists, so they'll get their colors from their respective tiers
    player_info = {}
    for name in tier_data['menA']:
        player_info[name] = {'gender': 'men', 'tier': 'A'}
    for name in tier_data['menB']:
        player_info[name] = {'gender': 'men', 'tier': 'B'}
    for name in tier_data['menC']:
        player_info[name] = {'gender': 'men', 'tier': 'C'}
    for name in tier_data['womenA']:
        player_info[name] = {'gender': 'women', 'tier': 'A'}
    for name in tier_data['womenB']:
        player_info[name] = {'gender': 'women', 'tier': 'B'}
    for name in tier_data['wildcard']:
        player_info[name] = {'gender': 'wildcard', 'tier': 'W'}
    
    # Ensure all setters are in player_info (they should already be in tier lists)
    # This is a safety check in case a setter is listed but not in any tier
    for setter_name in tier_data['setter']:
        if setter_name not in player_info:
            # Default to wildcard if setter is not found in any tier
            player_info[setter_name] = {'gender': 'wildcard', 'tier': 'W'}
    
    # Generate at least 10 groups (20 teams minimum)
    min_groups = 10
    teams_per_batch = 4  # Create 4 teams per batch, then reset setters
    all_teams = []
    
    # Calculate how many batches we need (each batch = 4 teams = 2 groups)
    batches_needed = math.ceil(min_groups / 2)  # 2 groups per batch

    for batch_num in range(batches_needed):
        # Create copies for this batch - setters are reset every 4 teams
        menA = tier_data['menA'].copy()
        menB = tier_data['menB'].copy()
        menC = tier_data['menC'].copy()
        womenA = tier_data['womenA'].copy()
        womenB = tier_data['womenB'].copy()
        wildcard = tier_data['wildcard'].copy()
        setter = tier_data['setter'].copy()

        # Create Tier objects with new point system
        # Men A - 4pt, Men B - 3pt, Women A - 3pt, Men C - 2pt, Women B - 2pt, Wildcard - 1pt
        TierMenA = Tier(menA, 4)
        TierMenB = Tier(menB, 3)
        TierWomenA = Tier(womenA, 3)
        TierMenC = Tier(menC, 2)
        TierWomenB = Tier(womenB, 2)
        TierWildcard = Tier(wildcard, 1)

        # Create players dictionary for easier management
        players_dict = {
            'menA': TierMenA,
            'menB': TierMenB,
            'menC': TierMenC,
            'womenA': TierWomenA,
            'womenB': TierWomenB,
            'wildcard': TierWildcard
        }
        
        # Create 4 teams per batch
        num_teams = 4
        
        # Create balanced teams
        teams = create_balanced_teams(players_dict, setter, num_teams)
        all_teams.extend(teams)

    # Display teams
    for index, team in enumerate(all_teams):
        print(f"Team {index+1} : (points value at: {team.total_points})")
        print(team.members)
        print("\n")

    # Export teams to Excel with grouped formatting
    excel_file = 'Fixtures.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = "All Teams"
    
    # Define color scheme
    colors = {
        'menA': 'b0e0e6',      # Dark blue for Men A
        'menB': 'afeeee',      # Medium blue for Men B
        'menC': 'e0ffff',      # Light blue for Men C
        'womenA': 'ffd6dc',    # Dark pink for Women A
        'womenB': 'ffe6e9',    # Light pink for Women B
        'wildcard': 'E6E6FA'   # Lavender for Wildcard
    }
    
    def get_player_color(player_name):
        """Get the fill color for a player based on their gender and tier"""
        if player_name not in player_info:
            return None
        
        info = player_info[player_name]
        gender = info['gender']
        tier = info['tier']
        
        # if gender == 'men':
        #     if tier == 'A':
        #         return colors['menA']
        #     elif tier == 'B':
        #         return colors['menB']
        #     elif tier == 'C':
        #         return colors['menC']
        # elif gender == 'women':
        #     if tier == 'A':
        #         return colors['womenA']
        #     elif tier == 'B':
        #         return colors['womenB']
        # elif gender == 'wildcard':
        #     return colors['wildcard']
        
        return None
    
    def header_cell(mcells, h1, text, team_nums, row, col, team_data):
        """
        Create a merged header cell and add team data below it.
        
        Args:
            mcells: Cell range to merge (e.g., 'B5:D5')
            h1: Header cell reference (e.g., 'B5')
            text: Header text (e.g., 'Group 1')
            team_nums: List of team numbers for column headers
            row: Starting row for column headers
            col: Starting column
            team_data: List of teams to display
        """
        ws.merge_cells(mcells)
        ws[h1] = text
        ws[h1].font = Font(bold=True, size=12)
        ws[h1].alignment = Alignment(horizontal='center')

        # Add column headers with points
        # \n({team_data[i].total_points}pts)
        headers = [""] + [f"Team {num}" 
                         for i, num in enumerate(team_nums)]
        for count, header in enumerate(headers):
            cell = ws.cell(row=row, column=col + count)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        # Add team data below headers
        max_players = max(len(team.members) for team in team_data) if team_data else 0
        
        for player_idx in range(max_players):
            ws.cell(row=row + 1 + player_idx, column=col).value = f"{player_idx + 1}"
            ws.cell(row=row + 1 + player_idx, column=col).font = Font(bold=True)
            ws.cell(row=row + 1 + player_idx, column=col).alignment = Alignment(horizontal='center')
            
            for team_idx, team in enumerate(team_data):
                if player_idx < len(team.members):
                    cell = ws.cell(row=row + 1 + player_idx, column=col + 1 + team_idx)
                    player_name = team.members[player_idx]
                    cell.value = player_name
                    cell.alignment = Alignment(horizontal='center')
                    
                    # Apply color based on player tier and gender
                    color = get_player_color(player_name)
                    if color:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    
    # Group teams by pairs (2 teams per group) - original layout
    teams_per_group = 2
    groups = [all_teams[i:i + teams_per_group] for i in range(0, len(all_teams), teams_per_group)]
    
    max_players = max(len(team.members) for team in all_teams) if all_teams else 6
    row_spacing = max_players + 3  # Add space for header, column headers, and a blank row
    
    for group_idx, group in enumerate(groups):
        row = 2 + (group_idx // 2) * row_spacing
        col = 2 + (group_idx % 2) * 4
        group_num = group_idx + 1
        team_nums = [group_idx * teams_per_group + 1 + i for i in range(len(group))]
        
        cell_ref = f"{chr(64 + col)}{row}:{chr(64 + col + len(group))}{row}"
        header_ref = f"{chr(64 + col)}{row}"
        
        header_cell(cell_ref, header_ref, f"Group {group_num}", team_nums, row + 1, col, group)
    
    wb.save(excel_file)
    print(f"\nTeams exported to {excel_file}")

main()