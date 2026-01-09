from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Create a workbook and select active sheet
wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# Add data rows
data = [
    [[1, 150, 200], [2, 180, 210]],
    [[1, 180, 210], [2, 190, 220]]
]

def header_cell(mcells, h1, text, tcount, row, col, row_data):
    ws.merge_cells(mcells)
    ws[h1] = text
    ws[h1].font = Font(bold=True, size=12)
    ws[h1].alignment = Alignment(horizontal='center')

    # Add column headers
    headers = ["", f"team {tcount}", f"team {tcount + 1}"]
    for count, header in enumerate(headers):
        cell = ws.cell(row=row, column=col + count)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    # Add data below headers
    for idx, value in enumerate(row_data):
        for j, val in enumerate(value):
            cell = ws.cell(row=row + 1 + idx, column=col + j)
            cell.value = val
            cell.alignment = Alignment(horizontal='center')

for i in range(2):
    row = 5 + (i // 2) * 9
    col = 2 + (i % 2) * 4
    group_num = i + 1
    tcount = i * 2 + 1
    
    cell_ref = f"{chr(64 + col)}{row}:{chr(64 + col + 2)}{row}"
    header_ref = f"{chr(64 + col)}{row}"
    
    header_cell(cell_ref, header_ref, f"Group {group_num}", tcount, row+1, col, data[i])





# for row_idx, row_data in enumerate(data, 4):
#     for col_idx, value in enumerate(row_data, 1):
#         ws.cell(row=row_idx, column=col_idx).value = value

# # Merge cells for a total row
# ws.merge_cells('A8:A9')
# ws['A8'] = "Total"
# ws['A8'].font = Font(bold=True)

# Save the workbook
wb.save('output.xlsx')